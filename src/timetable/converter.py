import openpyxl
import json
import sys
import os
from typing import Dict, List

# Mapping
day_map = {"จันทร์": "Mon", "อังคาร": "Tue", "พุธ": "Wed", "พฤหัสบดี": "Thu", "ศุกร์": "Fri"}

subject_map = {
    "คณิตศาสตร์": "Math",
    "วิทยาศาสตร์": "Science",
    "ภาษาอังกฤษ": "English",
    "ภาษาไทย": "Thai",
    "ประวัติศาสตร์": "History",
    "สังคมศึกษา": "Social",
    "ศิลปะ": "Art",
    "การงานอาชีพ": "Occupation",
    "สุขศึกษาและพลศึกษา": "Health Ed",
    "สุขศึกษาฯ(พละ)": "Physical Ed",
    "กิจกรรมพัฒนาผู้เรียน": "Activity",
    "ชุมนุม": "Club",
    # Added from unknown subjects
    "การงาน": "Occupation",
    "การป้องกันการทุจริต": "Anti-Corruption",
    "คณิตประยุกต์": "Applied Math",
    "คอมพิวเตอร์": "Computer",
    "ดนตรี-นาฏศิลป์": "Music-Drama",
    "ทัศนศิลป์": "Visual Arts",
    "ประว้ติศาสตร์": "History", # Typo from excel
    "ป้องกันการทุจริต": "Anti-Corruption",
    "ป้องการการทุจริต": "Anti-Corruption",
    "ภาษาอังกฤษเพิ่มเติม": "English Extra",
    "ภาษาไทยเพิ่มเติม": "Thai Extra",
    "วิทยาการคำนวณ": "Computer Science",
    "วิทยาศาสตร์เพิ่ม ฯ": "Science Extra",
    "วิทยาศาสตร์แบบสะเต็มศึกษา": "STEM Education",
    "ศิลปะ(ดนตรี)": "Art (Music)",
    "ศิลปะ(ทัศนศิลป์)": "Art (Visual)",
    "สุขศึกษาฯ": "Health Ed",
    "สุขศึกษาฯ (พละ)": "Physical Ed",
}

teacher_map = {
    "ครูสุกฤษฎิ์": "T001",
    "ครูอำพร": "T002",
    "ครูกฤตชยากร": "T003",
    "ครูพิมล": "T004",
    "ครูสุจิตร": "T005",
    "ครูปาณิสรา": "T006",
    "ครูวิยะดา": "T007",
    "ครูดวงใจ": "T008",
    "ครูสุขุมาภรณ์": "T009",
    "ครูพัฒนศักดิ์": "T010",
    "ครูบัวลอย": "T011",
    "ครูอภิชญา": "T012",
    "ครูสรัญญา": "T013",
    "ครูณัฏฐเศรษฐาวิชญ์": "T014",
    "ครูจุฑารัตน์": "T015",
    "ครูจิตยาภรณ์": "T016",
    "ครูจรรยาภรณ์": "T017",
    "ครูสิทธิศักดิ์": "T018"
}

def convert_timetable(file_path: str, output_file: str = "timetable_output.json") -> List[Dict]:
    """
    Convert Excel timetable to JSON format.

    Args:
        file_path: Path to the Excel file (.xlsm)
        output_file: Output JSON file path

    Returns:
        List of timetable entries

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        KeyError: If required worksheet is missing
    """
    # Validate input file
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    print(f"Loading workbook: {file_path}")
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        raise ValueError(f"Failed to load Excel file: {e}")

    # ชื่อ worksheet ที่ต้องการ
    sheet_names = ["ป1.-3 เทอม2 ปี2566", "ป. 4-6 เทอม2 ปี2566", "ม.1-3เทอม2ปี2566"]

    all_timetables = []
    unknown_subjects = set()
    unknown_teachers = set()

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found, skipping...")
            continue

        print(f"Processing sheet: {sheet_name}")
        ws = wb[sheet_name]

        # อ่าน period_id จาก row 2
        # Store tuples of (column_index, period_number) - skip non-numeric entries like lunch break
        # Handle both numeric periods (ม.1-3: "1", "2", "3") and time-based (ป.1-6: "09.00-10.00")
        periods = []
        period_counter = 1  # For mapping time-based periods to sequential numbers

        for col in range(3, ws.max_column + 1):
            period_value = ws.cell(row=2, column=col).value
            if period_value:
                period_str = str(period_value).strip()

                # Try to parse as integer first (for middle school: "1", "2", "3")
                try:
                    period_num = int(period_value)
                    periods.append((col, period_num))
                except (ValueError, TypeError):
                    # Check if it's a time range format (for elementary: "09.00-10.00")
                    if '-' in period_str and any(char.isdigit() for char in period_str):
                        # Valid time range - map to sequential period number
                        periods.append((col, period_counter))
                        period_counter += 1
                    # Otherwise skip (lunch break text, etc.)

        row = 3
        day = None
        class_id = None

        # Only process first table (up to row 32) to avoid duplicate data
        while row <= min(32, ws.max_row):
            day_val = ws.cell(row=row, column=1).value
            class_val = ws.cell(row=row, column=2).value

            # แปลง day/class (handle merged)
            day = day_map.get(day_val, day) if day_val else day
            class_id = class_val if class_val else class_id

            # loop แต่ละคาบ
            for col, period_num in periods:
                subject = ws.cell(row=row, column=col).value
                if subject and isinstance(subject, str):
                    subject = ''.join(char for char in subject.strip() if not char.isdigit())
                teacher = ws.cell(row=row + 1, column=col).value

                if subject and teacher:
                    teacher_stripped = teacher.strip()
                    subject_stripped = subject.strip()

                    # Track unknown mappings
                    if teacher_stripped not in teacher_map:
                        unknown_teachers.add(teacher_stripped)
                    if subject_stripped not in subject_map:
                        unknown_subjects.add(subject_stripped)

                    all_timetables.append({
                        "teacher_id": teacher_map.get(teacher_stripped, teacher_stripped),
                        "subject_id": subject_map.get(subject_stripped, subject_stripped),
                        "day_id": day,
                        "period_id": period_num,  # Use actual period number from Excel
                        "class_id": str(class_id)
                    })

            row += 2  # ข้าม 2 แถว (subject + teacher)

    # Report unknown mappings
    if unknown_subjects:
        print(f"\nWarning: {len(unknown_subjects)} unknown subject(s) found:")
        for subj in sorted(unknown_subjects):
            print(f"   - {subj}")

    if unknown_teachers:
        print(f"\nWarning: {len(unknown_teachers)} unknown teacher(s) found:")
        for teacher in sorted(unknown_teachers):
            print(f"   - {teacher}")

    # Close the workbook to release file handles
    wb.close()

    # Export JSON
    print(f"\nSaving to {output_file}...")
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(all_timetables, f, ensure_ascii=False, indent=2)

    print(f"Successfully saved {len(all_timetables)} timetable entries to {output_file}")
    return all_timetables


def main():
    """Main entry point for the script."""
    if len(sys.argv) < 2:
        print("Usage: python excel_converting.py <excel_file> [output_file]")
        print("\nExample:")
        print("  python excel_converting.py timetable.xlsm")
        print("  python excel_converting.py timetable.xlsm output.json")
        sys.exit(1)

    file_path = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else "timetable_output.json"

    try:
        convert_timetable(file_path, output_file)
    except Exception as e:
        print(f"\n❌ Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
