import openpyxl
import json

file_path = "your_timetable.xlsm"
wb = openpyxl.load_workbook(file_path)

# Mapping
day_map = {"จันทร์": "Mon", "อังคาร": "Tue", "พุธ": "Wed", "พฤหัสบดี": "Thu", "ศุกร์": "Fri"}

subject_map = {
    "คณิตศาสตร์": "Math",
    "วิทยาศาสตร์": "Science",
    "ภาษาอังกฤษ": "English",
    "ภาษาไทย": "Thai",
    "ประวัติศาสตร์": "History",
    "สังคมศึกษา": "Social",
    "ศิลปะ": "Art",
    "การงานอาชีพ": "Occupation",
    "สุขศึกษาและพลศึกษา": "Health Ed",
    "สุขศึกษาฯ(พละ)": "Physical Ed",
    "กิจกรรมพัฒนาผู้เรียน": "Activity",
    "ชุมนุม": "Club"
}

teacher_map = {
    "ครูสุกฤษฎิ์": "T001",
    "ครูอำพร": "T002",
    "ครูกฤตชยากร": "T003",
    "ครูพิมล": "T004",
    "ครูสุจิตร": "T005",
    "ครูปาณิสรา": "T006",
    "ครูวิยะดา": "T007",
    "ครูดวงใจ": "T008",
    "ครูสุขุมาภรณ์": "T009",
    "ครูพัฒนศักดิ์": "T010",
    "ครูบัวลอย": "T011",
    "ครูอภิชญา": "T012",
    "ครูสรัญญา": "T013",
    "ครูณัฏฐเศรษฐาวิชญ์": "T014",
    "ครูจุฑารัตน์": "T015",
    "ครูจิตยาภรณ์": "T016",
    "ครูจรรยาภรณ์": "T017",
    "ครูสิทธิศักดิ์": "T018"
}

# ชื่อ worksheet ที่ต้องการ
sheet_names = ["ป1.-3 เทอม2 ปี2566", "ป. 4-6 เทอม2 ปี2566", "ม.1-3เทอม2ปี2566"]

all_timetables = []

for sheet_name in sheet_names:
    ws = wb[sheet_name]

    # อ่าน period_id จาก row 2
    periods = []
    for col in range(3, ws.max_column + 1):
        if ws.cell(row=2, column=col).value:
            periods.append(col)

    timetable = []

    row = 3
    day = None
    class_id = None

    while row <= ws.max_row:
        day_val = ws.cell(row=row, column=1).value
        class_val = ws.cell(row=row, column=2).value

        # แปลง day/class (handle merged)
        day = day_map.get(day_val, day) if day_val else day
        class_id = class_val if class_val else class_id

        # loop แต่ละคาบ
        for idx, col in enumerate(periods, start=1):
            subject = ws.cell(row=row, column=col).value
            teacher = ws.cell(row=row + 1, column=col).value

            # print(f"-------->{subject_map.get(subject.strip(), "UNKNOWN")}")

            # import pdb
            # pdb.set_trace

            if subject and teacher:
                all_timetables.append({
                    "teacher_id": teacher_map.get(teacher.strip(), "UNKNOWN"),
                    "subject_id": subject_map.get(subject.strip(), "UNKNOWN"),
                    "day_id": day,
                    "period_id": idx,
                    "class_id": str(class_id)
                })

        row += 2  # ข้าม 2 แถว (subject + teacher)


# Export JSON
output_file = "timetable_output.json"
with open(output_file, "w", encoding="utf-8") as f:
    json.dump(all_timetables, f, ensure_ascii=False, indent=2)

print(f"JSON data has been saved to {output_file}")
