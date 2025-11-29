"""Check if T011 really has scheduling conflicts in Excel"""
import openpyxl

wb = openpyxl.load_workbook("ตารางเรียนเทอม2 ปี 68-2 .xlsm")
ws = wb["ม.1-3เทอม2ปี2566"]

# Map teacher name from row 35 to see who T011 is
teacher_map = {
    "ครูภัทราพร": "T001", "ครูอรอนงค์": "T002", "ครูสุนันท์": "T003",
    "ครูกัลยา": "T004", "ครูปิยฉัตร": "T005", "ครูสุมาลี": "T006",
    "ครูธนภัทร": "T007", "ครูไพรินทร์": "T008", "ครูสมชาย": "T009",
    "ครูสุรพงษ์": "T010", "ครูอัจฉรา": "T011", "ครูสุพัตรา": "T012",
}

# Find T011's actual name
# From actual teacher_map in excel_converting.py:36
t011_name = "ครูบัวลอย"

print(f"T011 is: {t011_name}")
print(f"\nSearching for {t011_name} in middle school sheet on Monday...")

# Get period column mappings
period_cols = {}
for col in range(3, ws.max_column + 1):
    period_value = ws.cell(row=2, column=col).value
    try:
        period_num = int(period_value)
        period_cols[period_num] = col
    except:
        pass

print(f"Period columns: {period_cols}")

# Search for T011's Monday classes
print(f"\nMonday classes taught by {t011_name}:")
row = 3
current_day = None
current_class = None

while row <= ws.max_row:
    day_val = ws.cell(row=row, column=1).value
    class_val = ws.cell(row=row, column=2).value

    if day_val:
        if "จันทร์" in str(day_val):
            current_day = "Mon"
        else:
            current_day = None

    if class_val:
        current_class = class_val

    if current_day == "Mon":
        for period_num, col in period_cols.items():
            subject = ws.cell(row=row, column=col).value
            teacher = ws.cell(row=row + 1, column=col).value

            if teacher and t011_name in str(teacher):
                print(f"  Class {current_class}, Period {period_num}, Column {col}: {subject} (teacher: {teacher})")

    row += 2

wb.close()
