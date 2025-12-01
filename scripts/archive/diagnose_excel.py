"""Diagnostic script to check Excel structure"""
import openpyxl

wb = openpyxl.load_workbook("ตารางเรียนเทอม2 ปี 68-2 .xlsm")
ws = wb["ม.1-3เทอม2ปี2566"]  # Check middle school sheet

print("Row 2 (Period numbers):")
for col in range(1, min(15, ws.max_column + 1)):
    val = ws.cell(row=2, column=col).value
    print(f"  Col {col}: {val}")

print("\nChecking columns with period data:")
periods = []
for col in range(3, ws.max_column + 1):
    period_num = ws.cell(row=2, column=col).value
    if period_num:
        periods.append((col, period_num))
        print(f"  Column {col} -> Period: {period_num}")

wb.close()
