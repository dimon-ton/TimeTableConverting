"""Check how elementary school periods are formatted"""
import openpyxl

wb = openpyxl.load_workbook("ตารางเรียนเทอม2 ปี 68-2 .xlsm")

# Check both elementary sheets
sheets = ["ป1.-3 เทอม2 ปี2566", "ป. 4-6 เทอม2 ปี2566"]

for sheet_name in sheets:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*70}")
        print(f"Sheet: {sheet_name}")
        print(f"{'='*70}")
        print("\nRow 2 (Period headers):")

        for col in range(1, min(15, ws.max_column + 1)):
            val = ws.cell(row=2, column=col).value
            print(f"  Col {col}: {val}")

        # Check what gets parsed with current logic
        print("\nWhat current parser would extract:")
        for col in range(3, ws.max_column + 1):
            period_value = ws.cell(row=2, column=col).value
            if period_value:
                try:
                    period_num = int(period_value)
                    print(f"  Column {col} -> Period {period_num} (numeric)")
                except (ValueError, TypeError):
                    print(f"  Column {col} -> SKIPPED: '{period_value}' (non-numeric)")

wb.close()
