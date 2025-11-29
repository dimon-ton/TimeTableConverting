"""Test the updated period parsing logic"""
import openpyxl

wb = openpyxl.load_workbook("ตารางเรียนเทอม2 ปี 68-2 .xlsm")

sheets = ["ป1.-3 เทอม2 ปี2566", "ป. 4-6 เทอม2 ปี2566", "ม.1-3เทอม2ปี2566"]

for sheet_name in sheets:
    if sheet_name not in wb.sheetnames:
        continue

    ws = wb[sheet_name]
    print(f"\n{'='*70}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*70}")

    # Simulate the new parsing logic
    periods = []
    period_counter = 1

    for col in range(3, ws.max_column + 1):
        period_value = ws.cell(row=2, column=col).value
        if period_value:
            period_str = str(period_value).strip()

            # Try to parse as integer first
            try:
                period_num = int(period_value)
                periods.append((col, period_num))
                print(f"  Col {col}: '{period_value}' -> Period {period_num} (numeric)")
            except (ValueError, TypeError):
                # Check if it's a time range format
                if '-' in period_str and any(char.isdigit() for char in period_str):
                    periods.append((col, period_counter))
                    print(f"  Col {col}: '{period_value}' -> Period {period_counter} (time-based)")
                    period_counter += 1
                else:
                    print(f"  Col {col}: '{period_value}' -> SKIPPED (not a period)")

    print(f"\nTotal periods extracted: {len(periods)}")

wb.close()
