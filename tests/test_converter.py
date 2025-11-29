"""
Unit tests for excel_converting.py

Run with: python test_excel_converting.py
"""

import unittest
import os
import json
import tempfile
from openpyxl import Workbook
from src.timetable.converter import convert_timetable, day_map, subject_map, teacher_map


class TestExcelConverting(unittest.TestCase):
    """Test cases for Excel timetable conversion"""

    def setUp(self):
        """Set up test data and temporary files"""
        self.temp_dir = tempfile.mkdtemp()
        self.test_excel_file = os.path.join(self.temp_dir, "test_timetable.xlsm")
        self.test_output_file = os.path.join(self.temp_dir, "test_output.json")

    def tearDown(self):
        """Clean up temporary files"""
        import time
        # Give some time for file handles to be released
        time.sleep(0.1)

        try:
            if os.path.exists(self.test_excel_file):
                os.remove(self.test_excel_file)
        except PermissionError:
            pass  # File is still locked, skip deletion

        try:
            if os.path.exists(self.test_output_file):
                os.remove(self.test_output_file)
        except PermissionError:
            pass

        try:
            if os.path.exists(self.temp_dir):
                os.rmdir(self.temp_dir)
        except (PermissionError, OSError):
            pass  # Directory not empty or still locked

    def create_mock_excel(self, include_all_sheets=True):
        """Create a mock Excel file for testing"""
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        sheet_names = ["ป1.-3 เทอม2 ปี2566", "ป. 4-6 เทอม2 ปี2566", "ม.1-3เทอม2ปี2566"]

        if include_all_sheets:
            sheets_to_create = sheet_names
        else:
            sheets_to_create = [sheet_names[0]]  # Only first sheet

        for sheet_name in sheets_to_create:
            ws = wb.create_sheet(sheet_name)

            # Header row
            ws['A1'] = 'วัน'
            ws['B1'] = 'ชั้น'
            ws['C1'] = 'คาบ'

            # Period numbers (row 2)
            for i in range(1, 6):  # 5 periods
                ws.cell(row=2, column=2+i, value=i)

            # Sample data: Monday, ป.1
            ws['A3'] = 'จันทร์'
            ws['B3'] = 'ป.1'
            ws['C3'] = 'คณิตศาสตร์'  # Math
            ws['D3'] = 'วิทยาศาสตร์'  # Science
            ws['C4'] = 'ครูสุกฤษฎิ์'  # T001
            ws['D4'] = 'ครูอำพร'  # T002

            # Sample data: Tuesday, ป.2 (testing merged cells behavior)
            ws['A5'] = ''  # Empty (merged cell)
            ws['B5'] = 'ป.2'
            ws['C5'] = 'ภาษาอังกฤษ'  # English
            ws['C6'] = 'ครูกฤตชยากร'  # T003

        wb.save(self.test_excel_file)
        wb.close()

    def test_file_not_found(self):
        """Test handling of non-existent Excel file"""
        with self.assertRaises(FileNotFoundError):
            convert_timetable("non_existent_file.xlsm", self.test_output_file)

    def test_basic_conversion(self):
        """Test basic timetable conversion"""
        self.create_mock_excel()

        result = convert_timetable(self.test_excel_file, self.test_output_file)

        # Should have created entries
        self.assertGreater(len(result), 0)

        # Check JSON file was created
        self.assertTrue(os.path.exists(self.test_output_file))

        # Verify JSON structure
        with open(self.test_output_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            self.assertIsInstance(data, list)

            # Check first entry structure
            if len(data) > 0:
                entry = data[0]
                self.assertIn('teacher_id', entry)
                self.assertIn('subject_id', entry)
                self.assertIn('day_id', entry)
                self.assertIn('period_id', entry)
                self.assertIn('class_id', entry)

    def test_day_mapping(self):
        """Test Thai to English day conversion"""
        self.create_mock_excel()
        result = convert_timetable(self.test_excel_file, self.test_output_file)

        # Check that days are correctly mapped
        days_found = set(entry['day_id'] for entry in result)

        # Should have 'Mon' from our mock data
        self.assertIn('Mon', days_found)

    def test_subject_mapping(self):
        """Test Thai to English subject conversion"""
        self.create_mock_excel()
        result = convert_timetable(self.test_excel_file, self.test_output_file)

        # Check that subjects are correctly mapped
        subjects_found = set(entry['subject_id'] for entry in result)

        # Our mock has Math, Science, English
        expected_subjects = {'Math', 'Science', 'English'}
        self.assertTrue(expected_subjects.issubset(subjects_found))

    def test_teacher_mapping(self):
        """Test Thai to English teacher ID conversion"""
        self.create_mock_excel()
        result = convert_timetable(self.test_excel_file, self.test_output_file)

        # Check that teachers are correctly mapped
        teachers_found = set(entry['teacher_id'] for entry in result)

        # Our mock has T001, T002, T003
        expected_teachers = {'T001', 'T002', 'T003'}
        self.assertTrue(expected_teachers.issubset(teachers_found))

    def test_period_numbering(self):
        """Test that period IDs are correctly assigned"""
        self.create_mock_excel()
        result = convert_timetable(self.test_excel_file, self.test_output_file)

        # Check that period_ids are positive integers
        for entry in result:
            self.assertIsInstance(entry['period_id'], int)
            self.assertGreater(entry['period_id'], 0)

    def test_merged_cells_handling(self):
        """Test handling of merged cells for day/class"""
        self.create_mock_excel()
        result = convert_timetable(self.test_excel_file, self.test_output_file)

        # All entries should have valid day_id and class_id
        for entry in result:
            self.assertIsNotNone(entry['day_id'])
            self.assertIsNotNone(entry['class_id'])
            self.assertNotEqual(entry['day_id'], '')
            self.assertNotEqual(entry['class_id'], '')

    def test_missing_sheet(self):
        """Test handling when some expected sheets are missing"""
        self.create_mock_excel(include_all_sheets=False)

        # Should still work with partial sheets
        result = convert_timetable(self.test_excel_file, self.test_output_file)

        # Should have some entries from the available sheet
        self.assertGreater(len(result), 0)

    def test_utf8_encoding(self):
        """Test that UTF-8 encoding is preserved in output"""
        self.create_mock_excel()
        convert_timetable(self.test_excel_file, self.test_output_file)

        # Read the JSON file and verify encoding
        with open(self.test_output_file, 'r', encoding='utf-8') as f:
            content = f.read()
            # Should be able to read without encoding errors
            self.assertIsInstance(content, str)

    def test_default_output_filename(self):
        """Test that default output filename works"""
        self.create_mock_excel()

        # Change to temp directory to avoid polluting current directory
        original_dir = os.getcwd()
        try:
            os.chdir(self.temp_dir)
            result = convert_timetable(self.test_excel_file)

            # Check default file was created
            default_file = os.path.join(self.temp_dir, "timetable_output.json")
            self.assertTrue(os.path.exists(default_file))

            # Clean up
            if os.path.exists(default_file):
                os.remove(default_file)
        finally:
            os.chdir(original_dir)

    def test_numeric_stripping_from_subjects(self):
        """Test that numeric characters are stripped from subject names"""
        # Create a special Excel with numbered subjects
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        ws = wb.create_sheet("ป1.-3 เทอม2 ปี2566")
        ws['A1'] = 'วัน'
        ws['B1'] = 'ชั้น'
        ws['C1'] = 'คาบ'
        ws.cell(row=2, column=3, value=1)

        ws['A3'] = 'จันทร์'
        ws['B3'] = 'ป.1'
        ws['C3'] = 'คณิตศาสตร์123'  # Math with numbers
        ws['C4'] = 'ครูสุกฤษฎิ์'

        wb.save(self.test_excel_file)
        wb.close()

        result = convert_timetable(self.test_excel_file, self.test_output_file)

        # Subject should be mapped correctly after stripping numbers
        self.assertGreater(len(result), 0)
        # The subject "คณิตศาสตร์123" should become "คณิตศาสตร์" and map to "Math"
        self.assertEqual(result[0]['subject_id'], 'Math')


class TestMappingDictionaries(unittest.TestCase):
    """Test cases for the mapping dictionaries"""

    def test_day_map_completeness(self):
        """Test that all weekdays are mapped"""
        expected_days = {'Mon', 'Tue', 'Wed', 'Thu', 'Fri'}
        mapped_days = set(day_map.values())
        self.assertEqual(mapped_days, expected_days)

    def test_subject_map_not_empty(self):
        """Test that subject map contains entries"""
        self.assertGreater(len(subject_map), 0)

        # Check some essential subjects exist
        thai_subjects = list(subject_map.keys())
        english_subjects = list(subject_map.values())

        self.assertIn('Math', english_subjects)
        self.assertIn('Science', english_subjects)
        self.assertIn('English', english_subjects)

    def test_teacher_map_not_empty(self):
        """Test that teacher map contains entries"""
        self.assertGreater(len(teacher_map), 0)

        # Check that teacher IDs follow the T### format
        for teacher_id in teacher_map.values():
            self.assertTrue(teacher_id.startswith('T'))
            self.assertTrue(teacher_id[1:].isdigit())


def run_tests():
    """Run all tests"""
    unittest.main(argv=[''], verbosity=2, exit=False)


if __name__ == "__main__":
    print("Running tests for excel_converting.py...")
    print("=" * 70)
    run_tests()
